import os, shutil
import smtplib
from email.header import Header
from email.mime.text import MIMEText

import win32com.client as win32

from openpyxl import load_workbook 
from faker import Faker

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOCUMENTS_DIR = os.path.join(BASE_DIR, 'documents')


def convert_xls_to_xlsx(ifname, ofname=None):
    """无损xls转xlsx，每次转一个文件"""
    if ofname is None:
        ofname = ifname + 'x'
        
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(ifname)

    wb.SaveAs(ofname, FileFormat = 51)  # 56 is for .xls extension
    wb.Close()
    excel.Application.Quit()


def batch_xls_to_xlsx():
    """无损xls转xlsx，批处理"""
    BACKUP_DIR = os.path.join(DOCUMENTS_DIR, 'backup')

    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

    ifnames = []
    for item in os.listdir(DOCUMENTS_DIR):
        if item.endswith('.xls'):
            file = os.path.join(DOCUMENTS_DIR, item)
            shutil.copy(file, BACKUP_DIR)
            ifnames.append(file)

    for ifname in ifnames:
        print('Converting...', ifname)
        convert_xls_to_xlsx(ifname)
        print('Done!')


def test_in_place_editing(fname='salary-12-07.xlsx'):
    """在原电子表格中修改值，不改变格式！"""
    in_place_file = os.path.join(DOCUMENTS_DIR, fname)
    if os.path.exists(in_place_file):
        print('Editing...', fname)

        myfake = Faker(locale='zh_CN')

        wb = load_workbook(in_place_file)
        ws = wb.active

        title = '2012年7月{}工资清单'.format(myfake.company_prefix())
        ws['A1'] = title

        for x in range(3, 105):
            _ = ws.cell(row=x, column=2, value=myfake.name())
            _ = ws.cell(row=x, column=3, value=myfake.email())
        wb.save(in_place_file)
        print('Done!')


def convert_sheet_to_xml():
    """sheet 转 xml"""
    excel_file = os.path.join(
        DOCUMENTS_DIR,
        'salary-12-07.xlsx'
    )
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    xml = f'<h3>{ws.title}</h3>'
    for row in ws.rows:
        if row[0].row == 1:
            # 处理表头
            xml += '<th>'
            for cell in row:
                xml += f'<tr>{cell.value}</tr>'
            xml += '</th>'
        else:
            # 处理表体
            xml += '<tr>'
            for cell in row:
                xml += f'<tr>{cell.value}</tr>'
            xml += '</tr>'
    return xml


def create_mails():
    """创建群邮件"""
    excel_file = os.path.join(
        DOCUMENTS_DIR,
        'salary-12-07.xlsx'
    )
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    mails = []
    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row-1,
                            values_only=True):
        m = {}
        m['receiver'] = row[2]
        message = f"""
            <h3>{row[1]}, 你好：</h3>
        """
        message += f"""
            {row[3:]}
        """
        m['message'] = message
        mails.append(m)
    return mails


if __name__ == '__main__':
    # 无损xls转xlsx，批处理
    # batch_xls_to_xlsx()
    # 测试在原电子表格中修改值
    # test_in_place_editing()
    # sheet 转 xml
    # print(convert_sheet_to_xml())
    # 创建群邮件
    print(create_mails())
